# -*- coding:utf-8 -*-
# 作   者：Ailian
# 开发时间：2021/10/22 下午2:18

import xlrd
import xlwt
import xlutils
import xlwings as xw
import xlsxwriter
import openpyxl
import pandas


xls_path = r'/Users/developer/Documents/系统相关导入导出模板/员工信息导入模板的副本.xls'
xlsx_path = r'/Users/developer/Documents/系统相关导入导出模板/员工信息导入模板的副本.xlsx'

# 3.1.1 xlrd 读取文件
xls = xlrd.open_workbook(xls_path)
xlsx = xlrd.open_workbook(xlsx_path)

xls_table = xls.sheets()[0]
xls_table_rownum = xls_table.nrows  # 统计该表中有效的总行数
xls_table_colnum = xls_table.ncols  # 统计该表中有效的总列数
xls_table_row = xls_table.row(2)  # 返回该行中所有的单元格 对象组成的列表
xls_table_row_slice = xls_table.row_slice(3)  # 返回该行中所有的单元格 对象组成的列表
xls_table_row_types = xls_table.row_types(3, start_colx=0, end_colx=None)  # 返回该行中所有单元格的数据类型组成的列表
# 注意：单元类型ctype：empty为0，string为1，number为2，date为3，boolean为4， error为5（左边为类型，右边为类型对应的值）
xls_table_row_value = xls_table.row_values(2, start_colx=0, end_colx=None)  # 返回该行中所有单元格的数据组成的列表
xls_table_row_len = xls_table.row_len(6)  # 返回该行中有效单元格长度

xls_table_col = xls_table.col(0)  # 返回该列中所有单元格 对象组成的列表
xls_table_col_slice = xls_table.col_slice(3)  # 返回该列中所有的单元格 对象组成的列表
xls_table_col_types = xls_table.col_types(3, start_rowx=1, end_rowx=None)  # 返回该列中所有单元格的数据类型组成的列表
# 注意：单元类型ctype：empty为0，string为1，number为2，date为3，boolean为4， error为5（左边为类型，右边为类型对应的值）
xls_table_col_value = xls_table.col_values(6, start_rowx=1, end_rowx=None)  # 返回该列中所有单元格的数据组成的列表


# # 3.1.2 xlwings 读取文件
# # xlwings 直接对接的是 apps，也就是 Excel应用程序，然后才是工作簿 books 和工作表 sheets，xlwings需要安装有 Excel 应用程序的环境 xlwings 可以读取 .xls 和 .xlsx 文件
# app = xw.App(visible=True, add_book=False, spec='wpsoffice')  # 程序可见，只打开不新建工作簿
# app.display_alerts = False  # 警告关闭
# app.screen_updating = False  # 屏幕更新关闭
# # wb = app.books.open(xls_path)
# wb = app.books.open(xlsx_path)
# # wb = xw.Book(xlsx_path)
# xlsx_table = wb.sheets["Sheet1"]  # 实例化工作表对象
# xlsx_table_value = xlsx_table.range("A2").value
# # wb.save()  # 保存文件
# # wb.close()  # 关闭文件
# # wb.quit()  # 关闭程序
#
# # wb_new = xw.Book()  # 创建一个新的工作簿
# # wb = xw.Book(xlsx_path)  # 连接到本地的现有文件

# 3.1.3 openpyxl 读取文件
# openpyxl 可以读取 .xlsx 文件
# wb = openpyxl.Workbook()  # 不需要有Excel文件存在，可以凭空产生一个
wb = openpyxl.load_workbook(xlsx_path)  # 需要传入一个Excel文件
wb_xlsx_table1 = wb.active  # 默认获取工作簿的第一张活动工作簿
# 注意：该函数调用工作表的索引（_active_sheet_index)，默认是0。除非你修改了这个值，否则使用改函数一直在是在对第一张表进行操作
# 想要获取别的 sheet 页，可以使用
wb_xlsx_table2 = wb["员工模板2"]

wb_xlsx_create_sheet1 = wb.create_sheet()  # 创建一个sheet，默认插在工作簿末尾
wb_xlsx_create_sheet2 = wb.create_sheet(0)  # 创建一个sheet，并且插在工作簿第一个位置

print(wb_xlsx_table2)
